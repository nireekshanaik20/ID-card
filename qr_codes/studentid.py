#This project is an AI-based Computer Vision system,
#  that uses pre-trained machine learning models for face detection and automated identity verification.”

import cv2
import pandas as pd
import tkinter as tk
from tkinter import messagebox
from pyzbar.pyzbar import decode, ZBarSymbol
from time import strftime as stf
import os

from time import strftime as stf
from openpyxl import load_workbook

# ---------------- LOAD DATA ----------------
try:
    df = pd.read_excel("students.xlsx")
    df["student_id"] = df["student_id"].astype(str).str.strip().str.upper()
except:
    messagebox.showerror("Error", "students.xlsx not found")
    exit()

# ---------------- FACE MODEL ----------------
face_cascade = cv2.CascadeClassifier(
    cv2.data.haarcascades + "haarcascade_frontalface_default.xml"
)

# ---------------- GOOGLE LENS STYLE BOX ----------------
def draw_corner_box(img, x1, y1, x2, y2, color=(255,0,0), thickness=4, length=45):
    # Top-left
    cv2.line(img, (x1, y1), (x1 + length, y1), color, thickness)
    cv2.line(img, (x1, y1), (x1, y1 + length), color, thickness)

    # Top-right
    cv2.line(img, (x2, y1), (x2 - length, y1), color, thickness)
    cv2.line(img, (x2, y1), (x2, y1 + length), color, thickness)

    # Bottom-left
    cv2.line(img, (x1, y2), (x1 + length, y2), color, thickness)
    cv2.line(img, (x1, y2), (x1, y2 - length), color, thickness)

    # Bottom-right
    cv2.line(img, (x2, y2), (x2 - length, y2), color, thickness)
    cv2.line(img, (x2, y2), (x2, y2 - length), color, thickness)

# ---------------- MAIN APP ----------------
root = tk.Tk()
root.title("Student Verification System")
root.geometry("700x620")
root.config(bg="#f2f2f2")

# ---------------- PAGE 1 : FACE SCAN ----------------
face_frame = tk.Frame(root, bg="#f2f2f2")
face_frame.pack(fill="both", expand=True)

tk.Label(face_frame, text="Face Verification",
         font=("Arial", 18, "bold"),
         bg="#f2f2f2").pack(pady=20)

tk.Label(face_frame, text="Align your FACE inside the box",
         font=("Arial", 12, "bold"),
         fg="green",
         bg="#f2f2f2").pack()

# ---------------- PAGE 2 : QR SCAN ----------------
qr_frame = tk.Frame(root, bg="#f2f2f2")

tk.Label(qr_frame, text="QR Code Verification",
         font=("Arial", 18, "bold"),
         bg="#f2f2f2").pack(pady=10)

result_frame = tk.Frame(qr_frame, bg="white", bd=2, relief="groove")
result_frame.pack(pady=10, fill="both", expand=True)

student_id_lbl = tk.Label(result_frame, text="Student ID:", font=("Arial", 12), bg="white")
name_lbl = tk.Label(result_frame, text="Name:", font=("Arial", 12), bg="white")
course_lbl = tk.Label(result_frame, text="Course:", font=("Arial", 12), bg="white")
email_lbl = tk.Label(result_frame, text="Email:", font=("Arial", 12), bg="white")
phone_lbl = tk.Label(result_frame, text="Phone:", font=("Arial", 12), bg="white")

for lbl in (student_id_lbl, name_lbl, course_lbl, email_lbl, phone_lbl):
    lbl.pack(anchor="w", padx=10, pady=5)

# ---------------- ATTENDANCE (EXCEL) ----------------



def att_reg(student_row):
    filename = "attendance.xlsx"
    sheet_name = stf("%Y-%m-%d")  # Each day = new sheet

    new_row = {
        "Student ID": student_row["student_id"],
        "Name": student_row["name"],
        "Course": student_row["course"],
        "Email": student_row["email"],
        "Phone": student_row["phone"],
        "Login Time": stf("%H:%M:%S")
    }

    # ---------- If file does NOT exist ----------
    if not os.path.exists(filename):
        df_att = pd.DataFrame([new_row])
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            df_att.to_excel(writer, sheet_name=sheet_name, index=False)
        return

    # ---------- If file exists ----------
    book = load_workbook(filename)

    if sheet_name in book.sheetnames:
        # Sheet exists → append
        df_existing = pd.read_excel(filename, sheet_name=sheet_name)
        df_updated = pd.concat(
            [df_existing, pd.DataFrame([new_row])],
            ignore_index=True
        )
    else:
        # New day → new sheet
        df_updated = pd.DataFrame([new_row])

    with pd.ExcelWriter(
        filename,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="replace"
    ) as writer:
        df_updated.to_excel(writer, sheet_name=sheet_name, index=False)


# ---------------- SHOW STUDENT ----------------
def fetch_student(sid):
    student = df[df["student_id"] == sid]
    if student.empty:
        messagebox.showerror("Error", "Student not found")
        return

    row = student.iloc[0].to_dict()

    student_id_lbl.config(text=f"Student ID: {row['student_id']}")
    name_lbl.config(text=f"Name: {row['name']}")
    course_lbl.config(text=f"Course: {row['course']}")
    email_lbl.config(text=f"Email: {row['email']}")
    phone_lbl.config(text=f"Phone: {row['phone']}")

    att_reg(row)

# ---------------- FACE SCAN AUTO ----------------
def auto_face_scan():
    cap = cv2.VideoCapture(0, cv2.CAP_DSHOW)

    while True:
        ret, frame = cap.read()
        if not ret:
            break

        h, w, _ = frame.shape
        size = 450
        x1, y1 = w//2 - size//2, h//2 - size//2
        x2, y2 = w//2 + size//2, h//2 + size//2

        roi = frame[y1:y2, x1:x2]
        gray = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)

        draw_corner_box(frame, x1, y1, x2, y2)

        faces = face_cascade.detectMultiScale(gray, 1.3, 5)
        if len(faces) > 0:
            cap.release()
            cv2.destroyAllWindows()
            face_frame.pack_forget()
            qr_frame.pack(fill="both", expand=True)
            auto_qr_scan()
            return

        cv2.imshow("Face Scan", frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()

# ---------------- QR SCAN AUTO ----------------
def auto_qr_scan():
    cap = cv2.VideoCapture(0, cv2.CAP_DSHOW)

    while True:
        ret, frame = cap.read()
        if not ret:
            break

        h, w, _ = frame.shape
        size = 450
        x1, y1 = w//2 - size//2, h//2 - size//2
        x2, y2 = w//2 + size//2, h//2 + size//2

        roi = frame[y1:y2, x1:x2]
        draw_corner_box(frame, x1, y1, x2, y2)

        decoded = decode(roi, symbols=[ZBarSymbol.QRCODE])
        for obj in decoded:
            sid = obj.data.decode().strip().upper()
            cap.release()
            cv2.destroyAllWindows()
            fetch_student(sid)
            return

        cv2.imshow("QR Scan", frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()

# ---------------- AUTO START ----------------
root.after(1000, auto_face_scan)
root.mainloop()
